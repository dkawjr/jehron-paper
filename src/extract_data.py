"""
extract_data.py
---------------
Reads the original Excel workbook (data/raw/Revised_Figures_and_Tables.xlsx)
and writes clean, tidy CSV files into data/ that every figure script consumes.

Run once:  py src/extract_data.py

Outputs
-------
data/positive_detections.csv   One row per positive GI-PCR detection (n=47):
                               day_post_sct, pathogen, category
data/table1_risk_factors.csv   Reconstructed 2x2 counts for unadjusted odds
                               ratios (positive vs negative GI-PCR).
data/table2_cox_model.csv      Adjusted Cox model hazard ratios (Table 2).
"""

from pathlib import Path
import csv
import openpyxl

ROOT = Path(__file__).resolve().parents[1]
RAW = ROOT / "data" / "raw" / "Revised_Figures_and_Tables.xlsx"
OUT = ROOT / "data"

# --- pathogen code -> name / category -------------------------------------
CODE2PATH = {
    0: "E. coli", 1: "Norovirus", 2: "Giardia", 3: "C. difficile",
    4: "Cryptosporidium", 5: "Adenovirus", 6: "Rotavirus", 7: "Sapovirus",
    8: "Salmonella", 9: "Yersinia", 10: "Other",
}
CATEGORY = {
    "E. coli": "Bacteria", "C. difficile": "Bacteria",
    "Salmonella": "Bacteria", "Yersinia": "Bacteria",
    "Norovirus": "Virus", "Adenovirus": "Virus",
    "Rotavirus": "Virus", "Sapovirus": "Virus",
    "Giardia": "Parasite", "Cryptosporidium": "Parasite",
    "Other": "Other",
}


def extract_detections(wb):
    """Column A = days post-SCT, Column B = pathogen code, for the 47 positives."""
    ws = wb["Figure 1"]
    rows = []
    for r in range(2, ws.max_row + 1):
        day = ws.cell(r, 1).value
        code = ws.cell(r, 2).value
        if day is None or code is None:
            continue
        try:
            day = int(day)
            code = int(code)
        except (TypeError, ValueError):
            continue
        if code not in CODE2PATH:
            continue
        path = CODE2PATH[code]
        rows.append({"day_post_sct": day,
                     "pathogen": path,
                     "category": CATEGORY[path]})
    rows.sort(key=lambda d: d["day_post_sct"])
    with open(OUT / "positive_detections.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["day_post_sct", "pathogen", "category"])
        w.writeheader()
        w.writerows(rows)
    print(f"positive_detections.csv  -> {len(rows)} rows")
    return rows


def write_table1():
    """
    Reconstructed 2x2 counts (positive vs negative GI-PCR) for unadjusted
    odds ratios. Counts are taken directly from Table 1 of the workbook.
    For each binary comparison we record the 'exposed' and 'reference' arms.
    Total positives = 41, total negatives = 49.
    """
    # variable, exposed_label, ref_label, exp_pos, exp_neg, ref_pos, ref_neg
    rows = [
        ("Male sex",              "Male",            "Female",         25, 20, 16, 29),
        ("No GVHD (<100d)",       "No GVHD",         "GVHD",           35, 30,  6, 19),
        ("Pre-colonization",      "Prior GI-PCR+",   "No prior GI-PCR",12,  9, 29, 40),
        ("Prior C. difficile",    "Prior C.diff+",   "No prior C.diff", 6,  7, 35, 42),
        ("Hypogammaglobulinemia", "IgG 200-700",     "Normal IgG",     12, 19, 29, 30),
        ("Neutropenia",           "Neutropenic",     "Normal ANC",      4,  7, 37, 42),
        ("Allograft",             "Allograft",       "Autograft",      20, 33, 20, 16),
        ("PPI use",               "PPI",             "No PPI",         24, 38, 17, 11),
        ("Chemotherapy + TBI",    "Chemo+TBI",       "Other cond.",     5, 14, 36, 35),
        ("Hospitalized",          "Hospitalized",    "Not hospitalized",21,36, 20, 13),
    ]
    with open(OUT / "table1_risk_factors.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["variable", "exposed", "reference",
                    "exp_pos", "exp_neg", "ref_pos", "ref_neg"])
        w.writerows(rows)
    print(f"table1_risk_factors.csv  -> {len(rows)} rows")


def write_table2():
    """
    Final (adjusted/reduced) multivariable Cox model, Table 2.
    HR with 95% CI. Reference levels included for readability.
    NOTE: the manuscript abstract garbles the sex estimate ('aHR 2.08,
    95% CI 0.26-0.90'); the correct adjusted estimate is Female HR 0.48
    (0.26-0.90), i.e. Male HR 2.08 relative to female reference.
    """
    rows = [
        # variable, level, hr, low, high, is_reference
        ("Sex",             "Male (ref)",       1.00, None, None, 1),
        ("Sex",             "Female",           0.48, 0.26, 0.90, 0),
        ("GVHD <100d",      "GVHD (ref)",       1.00, None, None, 1),
        ("GVHD <100d",      "No GVHD",          2.59, 1.12, 5.96, 0),
        ("Hospitalization", "Hospitalized (ref)",1.00, None, None, 1),
        ("Hospitalization", "Not hospitalized", 1.62, 0.86, 3.03, 0),
    ]
    with open(OUT / "table2_cox_model.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["variable", "level", "hr", "ci_low", "ci_high", "is_reference"])
        w.writerows(rows)
    print(f"table2_cox_model.csv     -> {len(rows)} rows")


def main():
    wb = openpyxl.load_workbook(RAW, data_only=True)
    extract_detections(wb)
    write_table1()
    write_table2()
    print("Done.")


if __name__ == "__main__":
    main()
